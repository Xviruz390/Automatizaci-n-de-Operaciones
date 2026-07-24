"""Convierte las operaciones de Reporte 1 en filas de Reporte 2.

Ejemplo:
  python automatizar_reporte.py --fecha 2026-07-23

Por seguridad genera "Reporte 2 automatizado.xlsx". Para reemplazar el archivo
original, añadir --in-place después de validar el resultado.
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import unicodedata
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "Reporte 1.xlsx"
TARGET = ROOT / "Reporte 2.xlsx"
CONFIG = ROOT / "configuracion_reporte.json"
LEGACY_OUTPUT = ROOT / "Reporte 2 automatizado corregido.xlsx"
MONTHS = (
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
)


def clean(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def key(value: object) -> str:
    return clean(value).lower()


def load_config() -> dict:
    with CONFIG.open(encoding="utf-8") as stream:
        return json.load(stream)


def aliases(config: dict, section: str, value: object, fallback: str) -> str:
    return config.get(section, {}).get(key(value), fallback if value else "")


def header_map(sheet) -> dict[str, int]:
    return {clean(cell.value): cell.column for cell in sheet[1] if cell.value is not None}


def find_column(headers: dict[str, int], names: tuple[str, ...]) -> int | None:
    for name in names:
        if clean(name) in headers:
            return headers[clean(name)]
    return None


def copy_row_style(sheet, source_row: int, target_row: int) -> None:
    for source, target in zip(sheet[source_row], sheet[target_row]):
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def parse_date(text: str) -> datetime:
    return datetime.strptime(text, "%Y-%m-%d")


def latest_operations_report() -> Path | None:
    reports: list[tuple[datetime, Path]] = []
    for report in ROOT.glob("Reporte Operaciones *.xlsx"):
        match = re.fullmatch(r"Reporte Operaciones (\d{2})-(\d{2})-(\d{2})\.xlsx", report.name)
        if match:
            day, month, year = (int(value) for value in match.groups())
            reports.append((datetime(2000 + year, month, day), report))
    if reports:
        return max(reports, key=lambda item: item[0])[1]
    if LEGACY_OUTPUT.exists():
        return LEGACY_OUTPUT
    return None


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--fecha", required=True, help="Fecha de las operaciones: AAAA-MM-DD")
    parser.add_argument("--canal", help="Canal de captación para las filas nuevas")
    parser.add_argument("--documento", help="Tipo de documento para las filas nuevas")
    parser.add_argument("--in-place", action="store_true", help="Sobrescribe Reporte 2.xlsx")
    args = parser.parse_args()

    operation_date = parse_date(args.fecha)
    config = load_config()
    source = load_workbook(SOURCE, data_only=False)
    output = ROOT / f"Reporte Operaciones {operation_date:%d-%m-%y}.xlsx"
    # En ejecuciones sucesivas, continuar desde el Reporte Operaciones más
    # reciente para conservar las operaciones de días anteriores.
    if args.in_place:
        target_path = TARGET
    elif output.exists():
        target_path = output
    else:
        latest = latest_operations_report() or TARGET
        shutil.copy2(latest, output)
        target_path = output
    destination = load_workbook(target_path)
    source_sheet = source.active
    destination_sheet = destination.active

    source_headers = header_map(source_sheet)
    columns = {
        "type": find_column(source_headers, ("Tipo de operación", "Tipo de operacion", "Operación")),
        "company": find_column(source_headers, ("Empresa", "Cliente")),
        "amount": find_column(source_headers, ("Cantidad", "Monto", "Importe")),
        "rate": find_column(source_headers, ("TC", "Tipo de cambio")),
        "bank": find_column(source_headers, ("Banco", "Entidad financiera")),
        "executive": find_column(source_headers, ("Ejecutivo", "Trader")),
    }
    missing = [name for name, column in columns.items() if column is None]
    if missing:
        raise ValueError(f"No se encontraron columnas en Reporte 1: {', '.join(missing)}")

    destination_headers = header_map(destination_sheet)
    required = [
        "N°", "FECHA", "PERIODO", "N° DE BOLETA", "TIPO DE DOCUMENTO",
        "NUMERO DE DOCUMENTO", "CLIENTE", "TIPO DE OPERACIÓN", "CANTIDAD",
        "TIPO DE CA", "BANCO", "EJECUTIVO", "CANAL DE CAPTACION", "TIPO DE CLIENTE",
    ]
    missing_destination = [name for name in required if clean(name) not in destination_headers]
    if missing_destination:
        raise ValueError(f"Faltan columnas en Reporte 2: {', '.join(missing_destination)}")

    # Reutiliza el tipo y número de documento ya registrados para cada cliente.
    document_by_client: dict[str, tuple[object, object]] = {}
    channel_by_client: dict[str, object] = {}
    client_column = destination_headers[clean("CLIENTE")]
    document_type_column = destination_headers[clean("TIPO DE DOCUMENTO")]
    document_number_column = destination_headers[clean("NUMERO DE DOCUMENTO")]
    channel_column = destination_headers[clean("CANAL DE CAPTACION")]
    for row in destination_sheet.iter_rows(min_row=2, values_only=True):
        client = row[client_column - 1]
        if client:
            document_by_client[key(client)] = (
                row[document_type_column - 1],
                row[document_number_column - 1],
            )
            if row[channel_column - 1]:
                channel_by_client[key(client)] = row[channel_column - 1]

    # Firma de duplicado: permite ejecutar el proceso varias veces sin repetir filas.
    existing = set()
    for row in destination_sheet.iter_rows(min_row=2, values_only=True):
        existing.add(tuple(key(row[destination_headers[clean(name)] - 1]) for name in (
            "FECHA", "CLIENTE", "TIPO DE OPERACIÓN", "CANTIDAD", "TIPO DE CA", "BANCO", "EJECUTIVO"
        )))

    number_column = destination_headers[clean("N°")]
    next_number = max((int(row[number_column - 1]) for row in destination_sheet.iter_rows(min_row=2, values_only=True)
                       if isinstance(row[number_column - 1], (int, float))), default=0) + 1
    series_by_company = config.get("serie_por_empresa", {})
    next_document_by_series: dict[str, int] = {}
    boleta_column = destination_headers[clean("N° DE BOLETA")]
    for row in destination_sheet.iter_rows(min_row=2, values_only=True):
        document = row[boleta_column - 1]
        match = re.fullmatch(r"([A-Z]\d{3})\s*[- ]\s*(\d+)", clean(document)) if document else None
        if match:
            series, number = match.groups()
            next_document_by_series[series] = max(next_document_by_series.get(series, 0), int(number) + 1)
    added = 0
    warnings: list[str] = []

    for row in range(2, source_sheet.max_row + 1):
        operation = source_sheet.cell(row, columns["type"]).value
        company_raw = source_sheet.cell(row, columns["company"]).value
        amount = source_sheet.cell(row, columns["amount"]).value
        rate = source_sheet.cell(row, columns["rate"]).value
        bank_raw = source_sheet.cell(row, columns["bank"]).value
        executive_raw = source_sheet.cell(row, columns["executive"]).value
        if clean(operation) not in {"COMPRA", "VENTA"} or not company_raw or not isinstance(amount, (int, float)):
            continue

        company = aliases(config, "empresas", company_raw, clean(company_raw))
        executive = aliases(config, "ejecutivos", executive_raw, clean(executive_raw))
        bank = aliases(config, "bancos", bank_raw, clean(bank_raw))
        operation = clean(operation)
        signature = (key(operation_date), key(company), key(operation), key(amount), key(rate), key(bank), key(executive))
        if signature in existing:
            continue

        known_document = document_by_client.get(key(company))
        if known_document and known_document[0] and known_document[1]:
            document_type, document_number = known_document
        else:
            companies = config.get("ruc_por_empresa", {})
            is_company = company in companies or company.endswith((" SAC", " SA", " EIRL", " SRL"))
            document_type = "RUC" if is_company else "DNI"
            document_number = companies.get(company, "PENDIENTE")
            warnings.append(f"Documento no encontrado en Reporte 2 para {company}; revisar {document_type}.")

        channel = channel_by_client.get(
            key(company), args.canal or config.get("canal_por_defecto", "PENDIENTE")
        )
        document_series = series_by_company.get(company, "B003")
        document_number_sequence = next_document_by_series.get(document_series, 1)

        new_row = destination_sheet.max_row + 1
        copy_row_style(destination_sheet, min(2, new_row), new_row)
        values = {
            "N°": next_number,
            "FECHA": operation_date,
            "PERIODO": MONTHS[operation_date.month - 1],
            "N° DE BOLETA": f"{document_series} - {document_number_sequence:03d}",
            "TIPO DE DOCUMENTO": document_type,
            "NUMERO DE DOCUMENTO": document_number,
            "CLIENTE": company,
            "TIPO DE OPERACIÓN": operation,
            "CANTIDAD": amount,
            "TIPO DE CA": rate,
            "BANCO": bank,
            "EJECUTIVO": executive,
            "CANAL DE CAPTACION": channel,
        }
        for name, value in values.items():
            destination_sheet.cell(new_row, destination_headers[clean(name)], value)
        client_cell = destination_sheet.cell(new_row, destination_headers[clean("CLIENTE")]).coordinate
        type_cell = destination_sheet.cell(new_row, destination_headers[clean("TIPO DE CLIENTE")]).coordinate
        providers = config.get("tipo_cliente_proveedor", [])
        # La fórmula evita depender de la posición de la fila y deja visible la regla.
        provider_list = ",".join(f'"{item}"' for item in providers)
        destination_sheet.cell(new_row, destination_headers[clean("TIPO DE CLIENTE")],
                               f'=IF(ISNUMBER(MATCH({client_cell},{{{provider_list}}},0)),"PROVEEDOR","CLIENTE")')
        existing.add(signature)
        next_number += 1
        next_document_by_series[document_series] = document_number_sequence + 1
        added += 1
        if clean(company_raw) not in {clean(k) for k in config.get("empresas", {})}:
            warnings.append(f"Empresa sin alias: {company_raw!r} -> {company}")
        if clean(executive_raw) not in {clean(k) for k in config.get("ejecutivos", {})}:
            warnings.append(f"Ejecutivo sin alias: {executive_raw!r} -> {executive}")

    destination.save(target_path)
    print(f"Filas agregadas: {added}")
    print(f"Archivo generado: {target_path.name}")
    for warning in sorted(set(warnings)):
        print(f"AVISO: {warning}")


if __name__ == "__main__":
    main()
